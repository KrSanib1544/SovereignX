# scripts/generate_demo_assets.py
"""
SOVEREIGN-X — Flagship Industrial Inspection Package Generator
Generates the 5 realistic synthetic demo engineering assets under demo/assets/:
1. inspection_report.pdf (Digital Vector PDF with ultrasonic thickness table)
2. scanned_report.pdf (Scanned Raster PDF with field dye-penetrant log)
3. equipment_photo.jpg (Macro image of casing weld seam crack)
4. maintenance_history.xlsx (5-year longitudinal Excel workbook)
5. maintenance_manual.pdf (OEM technical manual with Table 8.4 replacement limit)
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont
import fitz  # PyMuPDF

DEMO_DIR = Path(__file__).resolve().parent.parent / "demo" / "assets"


def ensure_demo_dir():
    DEMO_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[*] Demo assets directory initialized: {DEMO_DIR}")


def generate_inspection_report_pdf():
    pdf_path = DEMO_DIR / "inspection_report.pdf"
    doc = fitz.open()

    # Page 1: Executive Summary & NDT Data
    page = doc.new_page(width=595, height=842)  # A4 size

    # Header
    page.draw_rect(fitz.Rect(40, 40, 555, 90), color=(0.1, 0.2, 0.4), fill=(0.92, 0.95, 0.98))
    page.insert_text((55, 62), "NON-DESTRUCTIVE TESTING & INTEGRITY REPORT", fontsize=14, fontname="helv", color=(0.05, 0.15, 0.35))
    page.insert_text((55, 78), "HYDROCARBON REFLUX UNIT — ASSET INTEGRITY DIVISION", fontsize=9, fontname="helv", color=(0.3, 0.3, 0.3))

    # Metadata box
    page.insert_text((40, 115), "EQUIPMENT METADATA", fontsize=11, fontname="helv", color=(0.1, 0.1, 0.1))
    meta_text = (
        "Equipment Tag: P-303B\n"
        "Component: Reflux Pump 3B Casing Shell\n"
        "Service: Hydrocarbon Reflux Circulation (Flammable / Class 1 Div 1)\n"
        "Design Pressure: 32.0 bar | Operating Temperature: 185 deg C\n"
        "Inspection Technique: High-Resolution Ultrasonic Thickness Gauging (UT)\n"
        "Date of Inspection: 2026-08-15 | Lead Inspector: ID-8842 (Certified NDT Level III)"
    )
    page.insert_textbox(fitz.Rect(40, 125, 555, 200), meta_text, fontsize=9, fontname="helv")

    # Ultrasonic Thickness Data Table Header
    page.insert_text((40, 220), "SECTION 3.2 — ULTRASONIC WALL THICKNESS MEASUREMENTS", fontsize=11, fontname="helv", color=(0.1, 0.1, 0.1))

    # Draw table
    table_rect = fitz.Rect(40, 235, 555, 380)
    page.draw_rect(table_rect, color=(0.7, 0.7, 0.7))

    # Headers
    page.draw_rect(fitz.Rect(40, 235, 555, 260), fill=(0.2, 0.3, 0.5))
    page.insert_text((50, 252), "Casing Node", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((150, 252), "Nominal Baseline", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((270, 252), "Measured Thickness", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((400, 252), "Thinning Deficit", fontsize=9, fontname="helv", color=(1, 1, 1))
    page.insert_text((490, 252), "Status", fontsize=9, fontname="helv", color=(1, 1, 1))

    # Row data
    rows = [
        ("Node A-01 (Suction Nozzle)", "4.80 mm", "4.75 mm", "-0.05 mm", "NORMAL"),
        ("Node B-04 (Volute Upper)", "4.80 mm", "4.60 mm", "-0.20 mm", "NORMAL"),
        ("Node C-12 (Discharge Belly)", "4.80 mm", "3.42 mm", "-1.38 mm", "CRITICAL WEAR"),
        ("Node D-08 (Seal Housing)", "4.80 mm", "4.65 mm", "-0.15 mm", "NORMAL"),
        ("Node E-02 (Flange Neck)", "5.20 mm", "5.10 mm", "-0.10 mm", "NORMAL"),
    ]

    y = 280
    for node, nom, meas, loss, stat in rows:
        if "CRITICAL" in stat:
            page.draw_rect(fitz.Rect(42, y - 14, 553, y + 6), fill=(1.0, 0.9, 0.9))
            color = (0.8, 0.1, 0.1)
        else:
            color = (0.2, 0.2, 0.2)

        page.insert_text((50, y), node, fontsize=9, fontname="helv", color=color)
        page.insert_text((160, y), nom, fontsize=9, fontname="helv", color=color)
        page.insert_text((280, y), meas, fontsize=9, fontname="helv", color=color)
        page.insert_text((410, y), loss, fontsize=9, fontname="helv", color=color)
        page.insert_text((490, y), stat, fontsize=9, fontname="helv", color=color)
        y += 22

    # Narrative Finding
    narrative = (
        "CRITICAL FINDING NOTE:\n"
        "At Node C-12 (Discharge Belly Casing), ultrasonic wall thickness has degraded to 3.42 mm, "
        "representing a cumulative thinning of 1.38 mm from nominal baseline (4.80 mm). "
        "Accelerated erosion-corrosion appears localized in the lower volute quadrant. "
        "Immediate cross-referencing with OEM replacement manual is mandatory."
    )
    page.draw_rect(fitz.Rect(40, 400, 555, 480), color=(0.8, 0.2, 0.2), fill=(0.98, 0.94, 0.94))
    page.insert_textbox(fitz.Rect(50, 410, 545, 475), narrative, fontsize=9, fontname="helv", color=(0.6, 0.05, 0.05))

    doc.save(pdf_path, deflate=True, garbage=4)
    doc.close()
    print(f"  [+] Generated: {pdf_path}")


def generate_scanned_report_pdf():
    pdf_path = DEMO_DIR / "scanned_report.pdf"

    # Create a 300 DPI simulated scanned sheet
    width, height = 1200, 1600
    img = Image.new("RGB", (width, height), color=(248, 246, 240))  # Slightly aged paper tint
    draw = ImageDraw.Draw(img)

    # Add simulated paper noise / header
    draw.rectangle([50, 50, 1150, 150], fill=(230, 235, 242), outline=(100, 110, 130), width=2)
    draw.text((70, 70), "FIELD DYE-PENETRANT EXAMINATION LOG", fill=(20, 30, 60))
    draw.text((70, 105), "Unit: Hydrocarbon Reflux Plant | Technician ID: T-409 | Date: 2026-08-16", fill=(70, 80, 100))

    # Stamped box
    draw.rectangle([850, 65, 1130, 135], outline=(180, 30, 30), width=3)
    draw.text((870, 85), "OFFLINE FIELD NDT", fill=(180, 30, 30))
    draw.text((885, 105), "VERIFIED LOG", fill=(180, 30, 30))

    # Notes Body
    notes = [
        "EXAMINATION METHOD: Color Contrast Solvent Removable Liquid Penetrant (ASTM E165)",
        "TARGET COMPONENT: Reflux Pump 3B Lower Volute Weld Seam W-202",
        "SURFACE PREPARATION: Wire brushed and degreased with solvent cleaner.",
        "",
        "FIELD OBSERVATIONS & DEFECT INDICATION:",
        "1. Linear defect indication detected along longitudinal weld seam W-202.",
        "2. Surface breaking fatigue crack confirmed extending 48 mm along seam axis.",
        "3. Crack aperture measures approximately 1.4 mm at widest opening.",
        "4. Minor weld seam porosity and cluster pitting observed in surrounding heat-affected zone (HAZ).",
        "",
        "TECHNICIAN VERDICT:",
        "CRITICAL SEAM FRACTURE: Crack depth extends beyond allowable cosmetic grind limit.",
        "Recommends visual photo confirmation and immediate replacement of casing.",
    ]

    y = 200
    for line in notes:
        draw.text((80, y), line, fill=(30, 35, 45))
        y += 45

    # Stamp signature
    draw.rectangle([750, 800, 1100, 920], outline=(40, 80, 150), width=2)
    draw.text((770, 820), "CERTIFIED INSPECTOR SIGN-OFF", fill=(40, 80, 150))
    draw.text((770, 850), "Sign: J. Miller (Level II PT/UT)", fill=(20, 40, 90))
    draw.text((770, 880), "Status: REJECT — REPLACE REQUIRED", fill=(180, 20, 20))

    # Convert Image to Raster PDF Page
    doc = fitz.open()
    temp_img_path = DEMO_DIR / "temp_scanned.png"
    img.save(temp_img_path)

    page = doc.new_page(width=595, height=842)
    page.insert_image(fitz.Rect(0, 0, 595, 842), filename=str(temp_img_path))
    doc.save(pdf_path, deflate=True, garbage=4)
    doc.close()

    if temp_img_path.exists():
        temp_img_path.unlink()
    print(f"  [+] Generated: {pdf_path}")


def generate_equipment_photo():
    jpg_path = DEMO_DIR / "equipment_photo.jpg"

    # Create a realistic metallic inspection photograph (800x600)
    width, height = 800, 600
    img = Image.new("RGB", (width, height), color=(60, 65, 75))  # Cast steel dark gray
    draw = ImageDraw.Draw(img)

    # Weld seam textured band across middle
    draw.rectangle([0, 260, 800, 340], fill=(80, 85, 95))
    draw.line([(0, 260), (800, 260)], fill=(45, 50, 60), width=3)
    draw.line([(0, 340), (800, 340)], fill=(45, 50, 60), width=3)

    # Visible jagged fatigue crack along the weld seam
    crack_points = [
        (220, 300), (260, 296), (300, 304), (340, 298),
        (380, 305), (420, 297), (460, 303), (500, 300)
    ]
    for i in range(len(crack_points) - 1):
        draw.line([crack_points[i], crack_points[i + 1]], fill=(15, 15, 20), width=5)
        # Red penetrant dye highlighting the crack
        draw.line([(crack_points[i][0], crack_points[i][1] + 2), (crack_points[i + 1][0], crack_points[i + 1][1] + 2)], fill=(200, 30, 30), width=2)

    # Callout Annotation Overlay
    draw.rectangle([180, 180, 580, 250], fill=(20, 25, 35), outline=(0, 180, 220), width=2)
    draw.text((195, 195), "DEFECT: LONGITUDINAL FATIGUE CRACK", fill=(255, 80, 80))
    draw.text((195, 215), "Length: 48 mm | Aperture: 1.4 mm | Location: Seam W-202", fill=(0, 220, 255))
    draw.line([(360, 250), (360, 295)], fill=(0, 180, 220), width=2)

    # Scale bar in bottom right
    draw.rectangle([580, 520, 770, 570], fill=(20, 20, 25), outline=(150, 150, 150), width=1)
    draw.line([(595, 545), (755, 545)], fill=(255, 255, 255), width=3)
    draw.text((640, 550), "SCALE: 50 mm", fill=(255, 255, 255))

    img.save(jpg_path, quality=95)
    print(f"  [+] Generated: {jpg_path}")


def generate_maintenance_history_xlsx():
    xlsx_path = DEMO_DIR / "maintenance_history.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: Thickness_Log
    ws1 = wb.active
    ws1.title = "Thickness_Log"

    headers1 = ["Year", "Component", "Location", "Thickness_mm", "Inspection_Type", "Inspector"]
    ws1.append(headers1)

    data1 = [
        (2022, "Pump_3B_Casing", "Node_C12", 4.50, "Ultrasonic_Gauging", "NDT_01"),
        (2023, "Pump_3B_Casing", "Node_C12", 4.28, "Ultrasonic_Gauging", "NDT_04"),
        (2024, "Pump_3B_Casing", "Node_C12", 4.07, "Ultrasonic_Gauging", "NDT_02"),
        (2025, "Pump_3B_Casing", "Node_C12", 3.85, "Ultrasonic_Gauging", "NDT_05"),
        (2026, "Pump_3B_Casing", "Node_C12", 3.42, "Ultrasonic_Gauging", "NDT_08"),
    ]
    for row in data1:
        ws1.append(row)

    # Sheet 2: Vibration_Log
    ws2 = wb.create_sheet(title="Vibration_Log")
    headers2 = ["Quarter", "Asset_ID", "Overall_RMS_mm_s", "Peak_Velocity_mm_s", "Alarm_Limit"]
    ws2.append(headers2)
    data2 = [
        ("2025_Q1", "P-303B", 2.1, 3.4, 4.5),
        ("2025_Q2", "P-303B", 2.4, 3.8, 4.5),
        ("2025_Q3", "P-303B", 3.1, 4.2, 4.5),
        ("2025_Q4", "P-303B", 3.9, 5.1, 4.5),
        ("2026_Q1", "P-303B", 4.7, 6.2, 4.5),  # Alarm exceeded
    ]
    for row in data2:
        ws2.append(row)

    # Sheet 3: Maintenance_Events
    ws3 = wb.create_sheet(title="Maintenance_Events")
    headers3 = ["Event_ID", "Date", "Asset", "Maintenance_Scope", "Action_Taken"]
    ws3.append(headers3)
    data3 = [
        ("EV-101", "2023-04-12", "P-303B", "Mechanical Seal Flush", "Replaced Plan 53A barrier fluid"),
        ("EV-204", "2024-09-18", "P-303B", "Impeller Balancing", "Cleaned suction eye deposits"),
        ("EV-312", "2025-11-05", "P-303B", "Casing NDT Check", "Noted baseline thinning rate: 0.215 mm/yr"),
    ]
    for row in data3:
        ws3.append(row)

    wb.save(xlsx_path)
    print(f"  [+] Generated: {xlsx_path}")


def generate_maintenance_manual_pdf():
    pdf_path = DEMO_DIR / "maintenance_manual.pdf"
    doc = fitz.open()

    # Page 1: Title & Overview
    page1 = doc.new_page(width=595, height=842)
    page1.draw_rect(fitz.Rect(40, 40, 555, 120), color=(0.1, 0.3, 0.5), fill=(0.15, 0.25, 0.45))
    page1.insert_text((60, 75), "API 610 CENTRIFUGAL PROCESS PUMPS", fontsize=16, fontname="helv", color=(1, 1, 1))
    page1.insert_text((60, 95), "OEM TECHNICAL SPECIFICATION & MAINTENANCE MANUAL (REV 4.2)", fontsize=10, fontname="helv", color=(0.8, 0.9, 1))

    page1.insert_text((40, 150), "SECTION 8 — CASING INTEGRITY & WEAR TOLERANCES", fontsize=12, fontname="helv", color=(0.1, 0.1, 0.1))
    p1_text = (
        "8.1 Purpose and Scope:\n"
        "This section prescribes the minimum allowable wall thicknesses, corrosion allowances, "
        "and mandatory structural retirement criteria for Model SX-4000 series centrifugal pumps (including Reflux Pump 3B).\n\n"
        "8.2 Structural Minimum Wall Thickness Criteria:\n"
        "All pressure-retaining components are designed in accordance with ASME Section VIII and API 610 11th Edition standards. "
        "Continuous operation is predicated on maintaining a structural safety factor >= 2.5 against internal burst pressure."
    )
    page1.insert_textbox(fitz.Rect(40, 170, 555, 300), p1_text, fontsize=9, fontname="helv")

    # Table 8.4
    page1.insert_text((40, 320), "TABLE 8.4 — MANDATORY CASING SHELL REPLACEMENT LIMITS", fontsize=11, fontname="helv", color=(0.8, 0.1, 0.1))
    page1.draw_rect(fitz.Rect(40, 335, 555, 450), color=(0.6, 0.6, 0.6))
    page1.draw_rect(fitz.Rect(40, 335, 555, 360), fill=(0.2, 0.2, 0.3))
    page1.insert_text((50, 352), "Equipment Model / Asset", fontsize=9, fontname="helv", color=(1, 1, 1))
    page1.insert_text((220, 352), "Nominal Baseline", fontsize=9, fontname="helv", color=(1, 1, 1))
    page1.insert_text((350, 352), "Minimum Allowable (Limit)", fontsize=9, fontname="helv", color=(1, 1, 1))
    page1.insert_text((480, 352), "Action on Breach", fontsize=9, fontname="helv", color=(1, 1, 1))

    table_data = [
        ("SX-2000 Series (Water Feed)", "4.00 mm", "3.00 mm", "Scheduled Overhaul"),
        ("SX-4000 Series (Reflux Pump 3B)", "4.80 mm", "4.00 mm", "MANDATORY SHUTDOWN"),
        ("SX-6000 Series (High Pressure Gas)", "6.50 mm", "5.20 mm", "MANDATORY SHUTDOWN"),
    ]

    y = 380
    for model, nom, lim, act in table_data:
        if "Reflux Pump 3B" in model:
            page1.draw_rect(fitz.Rect(42, y - 14, 553, y + 6), fill=(1.0, 0.95, 0.9))
            color = (0.7, 0.1, 0.1)
        else:
            color = (0.2, 0.2, 0.2)
        page1.insert_text((50, y), model, fontsize=9, fontname="helv", color=color)
        page1.insert_text((230, y), nom, fontsize=9, fontname="helv", color=color)
        page1.insert_text((360, y), lim, fontsize=9, fontname="helv", color=color)
        page1.insert_text((480, y), act, fontsize=9, fontname="helv", color=color)
        y += 24

    # Mandatory Policy Clause
    clause_text = (
        "CRITICAL OPERATIONAL REQUIREMENT (TABLE 8.4 CLAUSE C):\n"
        "For Pump 3B (Model SX-4000), the absolute minimum allowable shell thickness is 4.00 mm. "
        "Operation with wall thickness below 4.00 mm (e.g. <= 3.99 mm) voids all manufacturer safety warranties "
        "and constitutes an uncontained containment breach hazard under operating reflux pressures. "
        "If inspection reveals wall thickness < 4.00 mm or any through-thickness fatigue cracking, "
        "the unit MUST BE DE-PRESSURIZED IMMEDIATELY AND REPLACED."
    )
    page1.draw_rect(fitz.Rect(40, 470, 555, 570), color=(0.8, 0.2, 0.2), fill=(1.0, 0.95, 0.95))
    page1.insert_textbox(fitz.Rect(50, 480, 545, 560), clause_text, fontsize=9, fontname="helv", color=(0.6, 0.05, 0.05))

    doc.save(pdf_path, deflate=True, garbage=4)
    doc.close()
    print(f"  [+] Generated: {pdf_path}")


def generate_all():
    print("=" * 80)
    print("  GENERATING SOVEREIGN-X FLAGSHIP INDUSTRIAL DEMO ASSETS")
    print("=" * 80)
    ensure_demo_dir()
    generate_inspection_report_pdf()
    generate_scanned_report_pdf()
    generate_equipment_photo()
    generate_maintenance_history_xlsx()
    generate_maintenance_manual_pdf()
    print("\n[SUCCESS] All 5 flagship industrial demo assets created in demo/assets/!\n")


if __name__ == "__main__":
    generate_all()
